import streamlit as st
import pandas as pd
import numpy as np
import io
import re
from datetime import datetime

st.set_page_config(
    page_title="PharmLogic Data QC Comparator",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Global CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

.hero {
  background: linear-gradient(135deg, #0f2044 0%, #1d4ed8 100%);
  border-radius: 16px; padding: 32px 40px 24px; margin-bottom: 28px; color: white;
}
.hero h1 { font-size: 1.9rem; font-weight: 700; margin: 0 0 6px; }
.hero p  { opacity: .8; margin: 0; font-size: .93rem; }

.tile-grid { display: flex; gap: 18px; flex-wrap: wrap; margin-bottom: 32px; }
.tile {
  flex: 1; min-width: 180px; max-width: 260px;
  background: #fff; border: 2px solid #e2e8f0; border-radius: 14px;
  padding: 24px 20px; text-align: center; cursor: pointer;
  transition: all .2s; box-shadow: 0 1px 4px rgba(0,0,0,.06);
}
.tile:hover  { border-color: #2563eb; box-shadow: 0 4px 16px rgba(37,99,235,.15); }
.tile.active { border-color: #2563eb; background: #eff6ff; }
.tile-icon   { font-size: 2rem; margin-bottom: 10px; }
.tile-label  { font-weight: 600; font-size: .95rem; color: #1e3a5f; }
.tile-sub    { font-size: .76rem; color: #64748b; margin-top: 4px; }

.badge-pass { background:#dcfce7; color:#166534; padding:2px 10px;
              border-radius:999px; font-size:.78rem; font-weight:600; }
.badge-fail { background:#fee2e2; color:#991b1b; padding:2px 10px;
              border-radius:999px; font-size:.78rem; font-weight:600; }
.badge-info { background:#dbeafe; color:#1e40af; padding:2px 10px;
              border-radius:999px; font-size:.78rem; font-weight:600; }
.badge-warn { background:#fef9c3; color:#854d0e; padding:2px 10px;
              border-radius:999px; font-size:.78rem; font-weight:600; }

.upload-label { font-weight:600; margin-bottom:6px; font-size:.93rem; color:#1e3a5f; }
.section-divider { border-top:2px solid #e2e8f0; margin:28px 0 20px; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

def load_file(uploaded) -> pd.DataFrame:
    raw = uploaded.read()
    name = uploaded.name.lower()
    # Excel
    if name.endswith((".xlsx", ".xls")):
        return pd.read_excel(io.BytesIO(raw), dtype=str)
    # Delimited
    for enc in ["utf-8", "latin-1", "cp1252"]:
        for sep in ["|", "\t", ",", ";"]:
            try:
                df = pd.read_csv(io.BytesIO(raw), sep=sep, dtype=str,
                                 encoding=enc, low_memory=False)
                if df.shape[1] > 3:
                    df.columns = df.columns.str.strip()
                    return df
            except Exception:
                pass
    raise ValueError(f"Cannot parse {uploaded.name}")

def blank(s: pd.Series) -> pd.Series:
    return s.isna() | s.astype(str).str.strip().isin(["", "nan", "NaN", "None", "NULL"])

def rb(ok) -> str:
    if ok is True:  return '<span class="badge-pass">✓ PASS</span>'
    if ok is False: return '<span class="badge-fail">✗ FAIL</span>'
    return '<span class="badge-info">ℹ INFO</span>'

def show_check(num, title, ok, msg, detail_df=None, extra=None):
    with st.expander(f"**{num}. {title}**  {rb(ok)}", expanded=(ok is False)):
        icon = "✅" if ok else ("❌" if ok is False else "ℹ️")
        st.markdown(f"{icon} {msg}")
        if extra:
            st.markdown(extra)
        if detail_df is not None and len(detail_df):
            st.dataframe(detail_df, use_container_width=True, hide_index=True)

def yyyymm_cols(df: pd.DataFrame) -> list:
    return [c for c in df.columns if re.fullmatch(r"\d{6}", c.strip())]

# ══════════════════════════════════════════════════════════════════════════════
# SUPER OAM CHECKS
# ══════════════════════════════════════════════════════════════════════════════

def soam_check1_ddd_id867(df1, df2):
    """DDD changes per ID_867 between versions."""
    id_col, ddd_col = "ID_867", "DDD"
    if id_col not in df1.columns or ddd_col not in df1.columns:
        return None, "ID_867 or DDD column missing", None
    def mapping(df):
        return (df[~blank(df[id_col]) & ~blank(df[ddd_col])]
                .groupby(id_col)[ddd_col]
                .agg(lambda x: " | ".join(sorted(x.unique())))
                .reset_index())
    m1 = mapping(df1).rename(columns={ddd_col: "DDD_v1"})
    m2 = mapping(df2).rename(columns={ddd_col: "DDD_v2"})
    merged = m1.merge(m2, on=id_col, how="outer")
    merged["DDD_v1"] = merged["DDD_v1"].fillna("—")
    merged["DDD_v2"] = merged["DDD_v2"].fillna("—")
    changed = merged[merged["DDD_v1"] != merged["DDD_v2"]].copy()
    ok = len(changed) == 0
    return ok, f"{len(changed)} ID_867(s) with DDD change between versions", changed if not ok else None

def soam_check2_exclusion_flags(df1, df2):
    """Subcat Excluded DDD?, PED Excluded?, Present In Override List? per DDD."""
    flag_cols = ["Subcat Excluded DDD?", "PED Excluded?", "Present In Override List?"]
    ddd_col = "DDD"
    available = [c for c in flag_cols if c in df1.columns and c in df2.columns]
    if not available or ddd_col not in df1.columns:
        return None, "Flag columns or DDD column missing", None

    def agg(df):
        sub = df[~blank(df[ddd_col])].copy()
        return (sub.groupby(ddd_col)[available]
                   .first()
                   .reset_index())

    a1 = agg(df1)
    a2 = agg(df2)
    merged = a1.merge(a2, on=ddd_col, how="outer", suffixes=("_Previous", "_Latest"))
    rows = []
    for _, r in merged.iterrows():
        ddd = r[ddd_col]
        for c in available:
            v1 = str(r.get(f"{c}_Previous", "")).strip()
            v2 = str(r.get(f"{c}_Latest", "")).strip()
            if v1 != v2:
                rows.append({"DDD": ddd, "Flag": c,
                             "Previous": v1, "Latest": v2})
    diff = pd.DataFrame(rows)
    ok = len(diff) == 0
    return ok, f"{len(diff)} flag change(s) across Exclusion columns", diff if not ok else None

def soam_check3_territory_unaligned(df1, df2):
    """Territory columns should not be entirely 'Unaligned'."""
    terr_cols = ["T1 ATP Alignment", "Current Alignment", "Zip to Terr Alignment"]
    findings = []
    for label, df in [("Previous", df1), ("Latest", df2)]:
        for c in terr_cols:
            if c not in df.columns:
                continue
            vals = df[c].dropna().str.strip().str.lower()
            total = len(vals)
            unaligned = (vals == "unaligned").sum()
            pct = round(unaligned / total * 100, 1) if total else 0
            if unaligned == total and total > 0:
                findings.append({"Version": label, "Column": c,
                                 "Status": "ALL Unaligned ⚠️",
                                 "Unaligned Count": unaligned, "Total": total, "%": pct})
            elif pct > 50:
                findings.append({"Version": label, "Column": c,
                                 "Status": f">{pct}% Unaligned",
                                 "Unaligned Count": unaligned, "Total": total, "%": pct})
    df_out = pd.DataFrame(findings)
    ok = len(df_out) == 0
    return ok, f"{len(df_out)} territory column(s) with high/full Unaligned", df_out if not ok else None

def soam_check4_smart_flags(df1, df2):
    """Compare SMART flags for all brands between versions."""
    smart_cols = [c for c in df1.columns if c.upper().startswith("SMART")]
    if not smart_cols:
        return None, "No SMART flag columns found", None
    ddd_col = "DDD"
    rows = []
    for c in smart_cols:
        if c not in df2.columns:
            rows.append({"Column": c, "Note": "Missing in Latest"})
            continue
        v1 = df1[ddd_col].value_counts() if ddd_col in df1.columns else None
        cnt_v1 = df1[c].value_counts().to_dict()
        cnt_v2 = df2[c].value_counts().to_dict()
        all_vals = set(list(cnt_v1.keys()) + list(cnt_v2.keys()))
        for val in all_vals:
            c1, c2 = cnt_v1.get(val, 0), cnt_v2.get(val, 0)
            if c1 != c2:
                rows.append({"SMART Column": c, "Flag Value": val,
                             "Count Previous": c1, "Count Latest": c2, "Delta": c2 - c1})
    diff = pd.DataFrame(rows)
    ok = len(diff) == 0
    return ok, f"{len(diff)} SMART flag distribution change(s)", diff if not ok else None

def soam_check5_child_parent(df1, df2):
    """Child/Parent ID and Name changes per DDD."""
    cols = ["DDD", "CHILD ID", "CHILD NAME", "PARENT ID", "PARENT NAME"]
    avail = [c for c in cols if c in df1.columns and c in df2.columns]
    if "DDD" not in avail:
        return None, "DDD or hierarchy columns missing", None
    def agg(df):
        return df[avail].drop_duplicates().dropna(subset=["DDD"])
    a1 = agg(df1)
    a2 = agg(df2)
    merged = a1.merge(a2, on="DDD", how="outer", suffixes=("_Previous", "_Latest"))
    diff_rows = []
    hier_cols = [c for c in avail if c != "DDD"]
    for _, r in merged.iterrows():
        changes = {}
        for c in hier_cols:
            v1 = str(r.get(f"{c}_Previous", "")).strip()
            v2 = str(r.get(f"{c}_Latest", "")).strip()
            if v1 != v2:
                changes[c] = {"Previous": v1, "Latest": v2}
        if changes:
            row = {"DDD": r["DDD"]}
            for c, chg in changes.items():
                row[f"{c} (Previous)"] = chg["Previous"]
                row[f"{c} (Latest)"]   = chg["Latest"]
            diff_rows.append(row)
    diff = pd.DataFrame(diff_rows)
    ok = len(diff) == 0
    return ok, f"{len(diff)} DDD(s) with Child/Parent changes", diff if not ok else None

def soam_check6_ic_channels(df1, df2):
    """Compare IC Channel flags for all brands between versions."""
    ic_cols = [c for c in df1.columns if c.upper().startswith("IC CHANNEL")]
    if not ic_cols:
        return None, "No IC Channel columns found", None
    rows = []
    for c in ic_cols:
        if c not in df2.columns:
            rows.append({"Column": c, "Note": "Missing in Latest"})
            continue
        cnt_v1 = df1[c].value_counts().to_dict()
        cnt_v2 = df2[c].value_counts().to_dict()
        all_vals = set(list(cnt_v1.keys()) + list(cnt_v2.keys()))
        for val in all_vals:
            c1, c2 = cnt_v1.get(val, 0), cnt_v2.get(val, 0)
            if c1 != c2:
                rows.append({"IC Channel Column": c, "Flag Value": val,
                             "Count Previous": c1, "Count Latest": c2, "Delta": c2 - c1})
    diff = pd.DataFrame(rows)
    ok = len(diff) == 0
    return ok, f"{len(diff)} IC Channel flag distribution change(s)", diff if not ok else None

def soam_check7_monthly_cols(df1, df2):
    """Compare yyyymm monthly columns present and their totals."""
    cols1 = set(yyyymm_cols(df1))
    cols2 = set(yyyymm_cols(df2))
    only_v1 = cols1 - cols2
    only_v2 = cols2 - cols1
    common  = cols1 & cols2
    rows = []
    for c in sorted(only_v1):
        rows.append({"Period": c, "Status": "Only in Previous",
                     "Sum Previous": df1[c].replace("", np.nan).astype(float, errors="ignore").sum(min_count=1),
                     "Sum Latest": "—"})
    for c in sorted(only_v2):
        rows.append({"Period": c, "Status": "Only in Latest",
                     "Sum Previous": "—",
                     "Sum Latest": df2[c].replace("", np.nan).astype(float, errors="ignore").sum(min_count=1)})
    for c in sorted(common):
        try:
            s1 = pd.to_numeric(df1[c], errors="coerce").sum()
            s2 = pd.to_numeric(df2[c], errors="coerce").sum()
            if abs(s1 - s2) > 0.01:
                rows.append({"Period": c, "Status": "Sum Changed",
                             "Sum Previous": round(s1, 2), "Sum Latest": round(s2, 2)})
        except Exception:
            pass
    diff = pd.DataFrame(rows) if rows else None
    ok = diff is None or len(diff) == 0
    msg = f"Monthly cols — Previous: {len(cols1)}, Latest: {len(cols2)}"
    if only_v1: msg += f" | Dropped: {', '.join(sorted(only_v1))}"
    if only_v2: msg += f" | Added: {', '.join(sorted(only_v2))}"
    return ok, msg, diff if not ok else None

# ══════════════════════════════════════════════════════════════════════════════
# NPS / IC FEED CHECKS  (fully dynamic — no hardcoded column names)
# ══════════════════════════════════════════════════════════════════════════════

def feed_check1_schema(df1, df2):
    """Column count, format (dtype inference), max-length per column."""
    rows = []
    all_cols = list(dict.fromkeys(list(df1.columns) + list(df2.columns)))
    for c in all_cols:
        in1 = c in df1.columns
        in2 = c in df2.columns
        if not in1:
            rows.append({"Column": c, "Issue": "Missing in Latest",
                         "Previous": "—", "Latest": "present"})
            continue
        if not in2:
            rows.append({"Column": c, "Issue": "Missing in Previous",
                         "Previous": "present", "Latest": "—"})
            continue
        # max length
        ml1 = df1[c].dropna().astype(str).str.len().max()
        ml2 = df2[c].dropna().astype(str).str.len().max()
        ml1 = 0 if (ml1 is None or (isinstance(ml1, float) and np.isnan(ml1))) else int(ml1)
        ml2 = 0 if (ml2 is None or (isinstance(ml2, float) and np.isnan(ml2))) else int(ml2)
        if ml1 != ml2:
            rows.append({"Column": c, "Issue": "Max length changed",
                         "Previous": ml1, "Latest": ml2})
    col_cnt_ok = df1.shape[1] == df2.shape[1]
    summary = (f"Column count — Previous: {df1.shape[1]}, Latest: {df2.shape[1]}"
               + (" ✅ Match" if col_cnt_ok else " ❌ Mismatch"))
    diff = pd.DataFrame(rows)
    ok = len(diff) == 0 and col_cnt_ok
    return ok, summary, diff if not ok else None

def feed_check2_totals(df1, df2):
    """Compare numeric column totals (catches NPS counts / factored qty etc.)."""
    num_cols = [c for c in df1.columns
                if c in df2.columns and
                pd.to_numeric(df1[c], errors="coerce").notna().sum() > len(df1) * 0.3]
    rows = []
    for c in num_cols:
        s1 = pd.to_numeric(df1[c], errors="coerce").sum()
        s2 = pd.to_numeric(df2[c], errors="coerce").sum()
        pct = abs(s1 - s2) / max(abs(s1), 1) * 100
        rows.append({"Column": c,
                     "Sum Previous": round(s1, 2),
                     "Sum Latest":   round(s2, 2),
                     "Delta":        round(s2 - s1, 2),
                     "Delta %":      round(pct, 2),
                     "Changed":      "Yes" if abs(s1 - s2) > 0.01 else "No"})
    df_out = pd.DataFrame(rows)
    changed = df_out[df_out["Changed"] == "Yes"] if len(df_out) else pd.DataFrame()
    ok = len(changed) == 0
    return ok, f"{len(changed)} numeric column(s) with total change", df_out if len(df_out) else None


def _infer_col_type(series: pd.Series) -> str:
    """Classify a column as Numeric, Date, or Text."""
    non_null = series.dropna().astype(str).str.strip()
    non_null = non_null[non_null != ""]
    if len(non_null) == 0:
        return "Empty"
    num_ratio = pd.to_numeric(non_null, errors="coerce").notna().mean()
    if num_ratio > 0.85:
        return "Numeric"
    try:
        pd.to_datetime(non_null.head(50), errors="raise")
        return "Date"
    except Exception:
        pass
    return "Text"


def feed_check5_col_type_length(df1, df2):
    """
    For every column compare:
      - inferred data type (Numeric / Date / Text / Empty)
      - max & avg value length
      - blank/null % in each version
      - if column is blank/empty in one version, flag whether the other is also blank
    """
    all_cols = list(dict.fromkeys(list(df1.columns) + list(df2.columns)))
    rows = []
    issues = []

    for c in all_cols:
        in1 = c in df1.columns
        in2 = c in df2.columns

        def stats(df, col):
            s = df[col]
            total = len(s)
            is_blank = blank(s)
            blank_cnt = int(is_blank.sum())
            blank_pct = round(blank_cnt / total * 100, 1) if total else 0
            non_null  = s[~is_blank].astype(str)
            max_len   = int(non_null.str.len().max()) if len(non_null) else 0
            avg_len   = round(non_null.str.len().mean(), 1) if len(non_null) else 0
            dtype     = _infer_col_type(s)
            return dtype, max_len, avg_len, blank_cnt, blank_pct, total

        if not in1:
            rows.append({"Column": c, "Issue": "Missing in Previous",
                         "Type Prev": "—", "Type Latest": "present",
                         "Max Len Prev": "—", "Max Len Latest": "—",
                         "Avg Len Prev": "—", "Avg Len Latest": "—",
                         "Blank% Prev": "—", "Blank% Latest": "—"})
            issues.append(c)
            continue
        if not in2:
            rows.append({"Column": c, "Issue": "Missing in Latest",
                         "Type Prev": "present", "Type Latest": "—",
                         "Max Len Prev": "—", "Max Len Latest": "—",
                         "Avg Len Prev": "—", "Avg Len Latest": "—",
                         "Blank% Prev": "—", "Blank% Latest": "—"})
            issues.append(c)
            continue

        t1, ml1, al1, bc1, bp1, tot1 = stats(df1, c)
        t2, ml2, al2, bc2, bp2, tot2 = stats(df2, c)

        col_issues = []
        if t1 != t2:
            col_issues.append(f"Type changed: {t1}→{t2}")
        if ml1 != ml2:
            col_issues.append(f"Max length changed: {ml1}→{ml2}")
        # Blank cross-check
        if t1 == "Empty" and t2 != "Empty":
            col_issues.append("Blank in Previous but NOT blank in Latest")
        elif t2 == "Empty" and t1 != "Empty":
            col_issues.append("Blank in Latest but NOT blank in Previous")
        elif t1 == "Empty" and t2 == "Empty":
            col_issues.append("Blank in BOTH versions")

        issue_str = " | ".join(col_issues) if col_issues else "OK"
        rows.append({
            "Column":         c,
            "Issue":          issue_str,
            "Type Prev":      t1,
            "Type Latest":    t2,
            "Max Len Prev":   ml1,
            "Max Len Latest": ml2,
            "Avg Len Prev":   al1,
            "Avg Len Latest": al2,
            "Blank% Prev":    bp1,
            "Blank% Latest":  bp2,
        })
        if col_issues:
            issues.append(c)

    df_out = pd.DataFrame(rows)
    ok = len(issues) == 0
    return ok, f"{len(issues)} column(s) with type/length/blank issues", df_out


def feed_check6_row_counts(df1, df2):
    """Compare total row counts and non-null row counts per column between versions."""
    rows_summary = [
        {"Metric": "Total Rows", "Previous": len(df1), "Latest": len(df2),
         "Delta": len(df2) - len(df1),
         "Status": "✅ Match" if len(df1) == len(df2) else "❌ Mismatch"},
    ]

    col_rows = []
    for c in df1.columns:
        if c not in df2.columns:
            continue
        cnt1 = int(df1[c].notna().sum() - blank(df1[c]).sum() +
                   df1[c].notna().sum() - df1[c].notna().sum())
        # simpler: non-blank count
        cnt1 = int((~blank(df1[c])).sum())
        cnt2 = int((~blank(df2[c])).sum())
        if cnt1 != cnt2:
            col_rows.append({
                "Column":              c,
                "Non-Blank Rows Prev": cnt1,
                "Non-Blank Rows Latest": cnt2,
                "Delta":               cnt2 - cnt1,
            })
    col_df = pd.DataFrame(col_rows)
    summary_df = pd.DataFrame(rows_summary)
    ok = (len(df1) == len(df2)) and len(col_df) == 0
    msg = (f"Total rows — Previous: {len(df1):,}, Latest: {len(df2):,}"
           + (" ✅ Match" if len(df1) == len(df2) else f" ❌ Delta: {len(df2)-len(df1):+,}")
           + f" | {len(col_df)} column(s) with non-blank row count change")
    return ok, msg, summary_df, col_df


def feed_check7_ic_totalrx_qty(df1, df2):
    """
    IC Feeds only: check TotalRx vs TotalQty equality (row-by-row count match)
    and compare row counts of each, not sums.
    """
    # Detect TotalRx and TotalQty columns
    def find_col(df, keywords):
        for c in df.columns:
            cu = c.upper().replace(" ", "").replace("_", "")
            if any(k in cu for k in keywords):
                return c
        return None

    rx_col1  = find_col(df1, ["TOTALRX",  "TOTAL_RX",  "TRX"])
    qty_col1 = find_col(df1, ["TOTALQTY", "TOTAL_QTY", "QTY"])
    rx_col2  = find_col(df2, ["TOTALRX",  "TOTAL_RX",  "TRX"])
    qty_col2 = find_col(df2, ["TOTALQTY", "TOTAL_QTY", "QTY"])

    rows = []
    issues = []

    for label, df, rxc, qtyc in [("Previous", df1, rx_col1, qty_col1),
                                   ("Latest",   df2, rx_col2, qty_col2)]:
        if not rxc:
            issues.append(f"TotalRx column not found in {label}")
            continue
        if not qtyc:
            issues.append(f"TotalQty column not found in {label}")
            continue

        rx_cnt  = int((~blank(df[rxc])).sum())
        qty_cnt = int((~blank(df[qtyc])).sum())
        rx_sum  = round(pd.to_numeric(df[rxc],  errors="coerce").sum(), 2)
        qty_sum = round(pd.to_numeric(df[qtyc], errors="coerce").sum(), 2)
        match   = "✅ Match" if rx_cnt == qty_cnt else "❌ Mismatch"
        rows.append({
            "Version":           label,
            "TotalRx Column":    rxc,
            "TotalQty Column":   qtyc,
            "TotalRx Row Count": rx_cnt,
            "TotalQty Row Count":qty_cnt,
            "Row Count Match":   match,
            "TotalRx Sum":       rx_sum,
            "TotalQty Sum":      qty_sum,
        })
        if rx_cnt != qty_cnt:
            issues.append(f"{label}: TotalRx rows ({rx_cnt:,}) ≠ TotalQty rows ({qty_cnt:,})")

    # Cross-version row count comparison
    if len(rows) == 2:
        for col_label, key in [("TotalRx", "TotalRx Row Count"),
                                ("TotalQty","TotalQty Row Count")]:
            v1, v2 = rows[0][key], rows[1][key]
            if v1 != v2:
                issues.append(f"{col_label} row count changed: Previous {v1:,} → Latest {v2:,}")

    df_out = pd.DataFrame(rows)
    ok = len(issues) == 0
    msg = " | ".join(issues) if issues else "TotalRx & TotalQty row counts match across both versions"
    return ok, msg, df_out if len(df_out) else None

def feed_check3_monthly(df1, df2):
    """Detect a period/date column and compare record presence month-by-month."""
    # Try to find a date/period column
    period_col = None
    for c in df1.columns:
        sample = df1[c].dropna().astype(str).head(100)
        # yyyymm pattern
        if sample.str.fullmatch(r"\d{6}").mean() > 0.7:
            period_col = c
            break
        # yyyy-mm-dd or mm/dd/yyyy
        try:
            pd.to_datetime(sample, errors="raise")
            period_col = c
            break
        except Exception:
            pass

    # Try to find a key column (HCP / account id)
    id_col = None
    for candidate in ["NPI", "HCP_ID", "ACCT_ID", "ACCOUNT_ID", "ID", "DEA",
                       "PRESCRIBER_ID", "PHYSICIAN_ID"]:
        if candidate in df1.columns:
            id_col = candidate
            break
    # fallback: first non-period column
    if not id_col:
        others = [c for c in df1.columns if c != period_col]
        id_col = others[0] if others else None

    if not period_col or not id_col:
        return None, "Could not auto-detect period or ID column for monthly comparison", None

    def per_period(df):
        df2c = df[[period_col, id_col]].copy()
        df2c[period_col] = df2c[period_col].astype(str).str.strip()
        return df2c.groupby(period_col)[id_col].apply(set).to_dict()

    p1 = per_period(df1)
    p2 = per_period(df2)
    all_periods = sorted(set(list(p1.keys()) + list(p2.keys())))
    rows = []
    for p in all_periods:
        s1, s2 = p1.get(p, set()), p2.get(p, set())
        added   = s2 - s1
        removed = s1 - s2
        rows.append({
            "Period":                p,
            "Count Previous":       len(s1),
            "Count Latest":         len(s2),
            "Added in Latest":      len(added),
            "Removed in Latest":    len(removed),
            "Sample Added":         ", ".join(list(added)[:5]) if added else "",
            "Sample Removed":       ", ".join(list(removed)[:5]) if removed else "",
        })
    df_out = pd.DataFrame(rows)
    changed = df_out[(df_out["Added in Latest"] > 0) | (df_out["Removed in Latest"] > 0)]
    ok = len(changed) == 0
    return ok, (f"Monthly comparison on '{period_col}' by '{id_col}' — "
                f"{len(changed)} period(s) with additions/removals"), df_out

def feed_check4_alignment_monthly(df1, df2, date_col, align_col, value_col):
    """
    Group by Alignment ID + yyyymm extracted from date_col, sum value_col.
    Compare Previous vs Latest and highlight changes.
    """
    def prep(df):
        d = df[[date_col, align_col, value_col]].copy()
        # Extract yyyymm — handles full dates (2024-01-15 → 202401) and bare yyyymm
        sample = d[date_col].dropna().astype(str).str.strip().head(50)
        if sample.str.fullmatch(r"\d{6}").mean() > 0.7:
            d["_period"] = d[date_col].astype(str).str.strip().str[:6]
        else:
            parsed = pd.to_datetime(d[date_col], errors="coerce")
            d["_period"] = parsed.dt.strftime("%Y%m")
        d[value_col] = pd.to_numeric(d[value_col], errors="coerce").fillna(0)
        return (d.dropna(subset=["_period", align_col])
                 .groupby(["_period", align_col], as_index=False)[value_col]
                 .sum()
                 .rename(columns={"_period": "Period",
                                  align_col: "Alignment_ID",
                                  value_col: "Value"}))

    g1 = prep(df1)
    g2 = prep(df2)

    # ── Merge for side-by-side comparison ────────────────────────────────────
    merged = g1.merge(g2, on=["Period", "Alignment_ID"], how="outer",
                      suffixes=("_Previous", "_Latest"))
    merged["Value_Previous"] = merged["Value_Previous"].fillna(0)
    merged["Value_Latest"]   = merged["Value_Latest"].fillna(0)
    merged["Delta"]          = merged["Value_Latest"] - merged["Value_Previous"]
    merged["Delta %"]        = (
        merged["Delta"] / merged["Value_Previous"].replace(0, np.nan) * 100
    ).round(2).fillna(0)
    merged["Status"] = merged.apply(
        lambda r: ("New in Latest"      if r["Value_Previous"] == 0 else
                   "Removed in Latest"  if r["Value_Latest"]   == 0 else
                   "Changed"            if abs(r["Delta"]) > 0.01 else
                   "Matched"), axis=1)

    # ── Monthly summary (period-level rollup) ─────────────────────────────────
    period_summary = (merged
        .groupby("Period")
        .agg(
            Alignments_Previous=("Value_Previous", lambda x: (x > 0).sum()),
            Alignments_Latest  =("Value_Latest",   lambda x: (x > 0).sum()),
            Sum_Previous       =("Value_Previous", "sum"),
            Sum_Latest         =("Value_Latest",   "sum"),
            New_Alignments     =("Status", lambda x: (x == "New in Latest").sum()),
            Removed_Alignments =("Status", lambda x: (x == "Removed in Latest").sum()),
            Changed_Alignments =("Status", lambda x: (x == "Changed").sum()),
        )
        .reset_index()
    )
    period_summary["Sum_Delta"] = (period_summary["Sum_Latest"]
                                   - period_summary["Sum_Previous"]).round(2)

    changed = merged[merged["Status"] != "Matched"]
    ok = len(changed) == 0
    msg = (f"Grouped by '{align_col}' + yyyymm from '{date_col}', "
           f"summing '{value_col}' — "
           f"{len(changed)} Alignment×Period combination(s) with changes")
    return ok, msg, period_summary, merged

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT (with QC Checklist tab)
# ══════════════════════════════════════════════════════════════════════════════

def build_excel(report_type: str, checks: list) -> bytes:
    """checks = list of {name, ok, msg, df}"""
    buf = io.BytesIO()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (PatternFill, Font, Alignment,
                                     Border, Side, numbers)
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to xlsxwriter-free openpyxl path
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            pd.DataFrame([{"Check": c["name"], "Status": "PASS" if c["ok"] else "FAIL",
                           "Detail": c["msg"]} for c in checks]).to_excel(
                w, sheet_name="Summary", index=False)
            for c in checks:
                if c.get("df") is not None and len(c["df"]):
                    sname = re.sub(r"[\\/*?:\[\]]", "_", c["name"])[:31]
                    c["df"].to_excel(w, sheet_name=sname, index=False)
        return buf.getvalue()

    wb = Workbook()
    wb.remove(wb.active)

    # Colour palette
    HDR_FILL  = PatternFill("solid", fgColor="0F2044")
    PASS_FILL = PatternFill("solid", fgColor="D1FAE5")
    FAIL_FILL = PatternFill("solid", fgColor="FEE2E2")
    INFO_FILL = PatternFill("solid", fgColor="DBEAFE")
    ALT_FILL  = PatternFill("solid", fgColor="F8FAFC")
    WHITE     = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_hdr(ws, row_num, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = HDR_FILL
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border = border

    def write_df(ws, df, start_row=1):
        if df is None or len(df) == 0:
            return
        for ri, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=start_row + ri - 1, column=ci, value=val)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                if ri == 1:
                    cell.fill = HDR_FILL
                    cell.font = Font(bold=True, color="FFFFFF", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.fill = ALT_FILL if ri % 2 == 0 else WHITE
                    cell.font = Font(size=9)
        # auto-width
        for col in ws.iter_cols(min_row=start_row, max_row=ws.max_row):
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)
        ws.row_dimensions[start_row].height = 30

    # ── 1. QC Checklist ──────────────────────────────────────────────────────
    ws_qc = wb.create_sheet("✅ QC Checklist")
    ws_qc.sheet_view.showGridLines = False
    ws_qc.column_dimensions["A"].width = 6
    ws_qc.column_dimensions["B"].width = 42
    ws_qc.column_dimensions["C"].width = 18
    ws_qc.column_dimensions["D"].width = 18
    ws_qc.column_dimensions["E"].width = 14
    ws_qc.column_dimensions["F"].width = 30

    title_cell = ws_qc.cell(1, 1,
        f"QC Checklist — {report_type}  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    title_cell.font = Font(bold=True, size=13, color="0F2044")
    title_cell.fill = PatternFill("solid", fgColor="EFF6FF")
    ws_qc.merge_cells("A1:F1")
    ws_qc.row_dimensions[1].height = 28

    hdr_row = ["#", "Check Description", "Auto Result", "Auto Finding",
               "Manual Sign-off", "QC Notes"]
    for ci, h in enumerate(hdr_row, 1):
        cell = ws_qc.cell(3, ci, h)
        cell.fill = HDR_FILL
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws_qc.row_dimensions[3].height = 24

    for ri, chk in enumerate(checks, 1):
        status = "PASS" if chk["ok"] is True else ("FAIL" if chk["ok"] is False else "INFO")
        fill = PASS_FILL if status == "PASS" else (FAIL_FILL if status == "FAIL" else INFO_FILL)
        row_data = [ri, chk["name"], status, chk["msg"][:120], "", ""]
        for ci, val in enumerate(row_data, 1):
            cell = ws_qc.cell(ri + 3, ci, val)
            cell.border = border
            cell.font = Font(size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if ci == 3:
                cell.fill = fill
                cell.font = Font(bold=True, size=9,
                    color="166534" if status == "PASS" else
                          ("991B1B" if status == "FAIL" else "1E40AF"))
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_qc.row_dimensions[ri + 3].height = 20

    # ── 2. Summary ────────────────────────────────────────────────────────────
    ws_s = wb.create_sheet("📋 Summary")
    ws_s.sheet_view.showGridLines = False
    summary_df = pd.DataFrame([
        {"#": i+1, "Check": c["name"],
         "Status": "PASS" if c["ok"] is True else ("FAIL" if c["ok"] is False else "INFO"),
         "Detail": c["msg"]}
        for i, c in enumerate(checks)
    ])
    write_df(ws_s, summary_df)
    # colour status column
    for row in ws_s.iter_rows(min_row=2, max_row=ws_s.max_row, min_col=3, max_col=3):
        for cell in row:
            if cell.value == "PASS":
                cell.fill = PASS_FILL
                cell.font = Font(bold=True, color="166534", size=9)
            elif cell.value == "FAIL":
                cell.fill = FAIL_FILL
                cell.font = Font(bold=True, color="991B1B", size=9)

    # ── 3. Detail sheets ──────────────────────────────────────────────────────
    for chk in checks:
        if chk.get("df") is not None and len(chk["df"]):
            sname = re.sub(r"[\\/*?:\[\]]", "_", chk["name"])[:31]
            ws_d = wb.create_sheet(sname)
            ws_d.sheet_view.showGridLines = False
            ws_d.cell(1, 1, chk["name"]).font = Font(bold=True, size=11, color="0F2044")
            ws_d.cell(2, 1, chk["msg"]).font  = Font(size=9, color="64748B")
            write_df(ws_d, chk["df"], start_row=4)

    wb.save(buf)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# UI — PAGE SELECTION
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<div class="hero">
  <h1>🔬 PharmLogic Data QC Comparator</h1>
  <p>Select a report type, upload <strong>Previous</strong> and <strong>Latest</strong> versions to run automated quality checks and produce a detailed comparison report.</p>
</div>
""", unsafe_allow_html=True)

REPORTS = {
    "OAM":       {"icon": "📊", "sub": "Outlet Account Master"},
    "Super OAM": {"icon": "🏢", "sub": "Extended OAM with Channels"},
    "NPS Feeds": {"icon": "💊", "sub": "Prescription Data Feed"},
    "IC Feeds":  {"icon": "⚡", "sub": "Incentive Comp Feed"},
}

if "report_type" not in st.session_state:
    st.session_state.report_type = None

st.markdown("### Step 1 — Select Report Type")

cols = st.columns(4)
for i, (name, meta) in enumerate(REPORTS.items()):
    with cols[i]:
        active = "active" if st.session_state.report_type == name else ""
        # Use button styled as tile
        if st.button(f"{meta['icon']}\n\n**{name}**\n\n*{meta['sub']}*",
                     key=f"tile_{name}",
                     use_container_width=True,
                     type="primary" if st.session_state.report_type == name else "secondary"):
            st.session_state.report_type = name
            st.rerun()

if not st.session_state.report_type:
    st.info("⬆️  Choose a report type above to continue.", icon="ℹ️")
    st.stop()

rtype = st.session_state.report_type
st.markdown(f"**Selected:** {REPORTS[rtype]['icon']} **{rtype}** — {REPORTS[rtype]['sub']}")
st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — FILE UPLOAD
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("### Step 2 — Upload Files")
col_a, col_b = st.columns(2)
with col_a:
    st.markdown('<div class="upload-label">📂 Previous Version</div>', unsafe_allow_html=True)
    f1 = st.file_uploader("Previous", type=["dat","csv","txt","xlsx","xls"],
                           key="f1", label_visibility="collapsed")
with col_b:
    st.markdown('<div class="upload-label">📂 Latest Version</div>', unsafe_allow_html=True)
    f2 = st.file_uploader("Latest", type=["dat","csv","txt","xlsx","xls"],
                           key="f2", label_visibility="collapsed")

if not f1 or not f2:
    st.info("⬆️  Upload both files to start the comparison.", icon="ℹ️")
    st.stop()

with st.spinner("Parsing files…"):
    try:
        df1 = load_file(f1)
        df2 = load_file(f2)
    except Exception as e:
        st.error(f"File parse error: {e}")
        st.stop()

c1, c2, c3, c4 = st.columns(4)
c1.metric("Previous Rows",    f"{df1.shape[0]:,}")
c2.metric("Previous Columns", df1.shape[1])
c3.metric("Latest Rows",      f"{df2.shape[0]:,}")
c4.metric("Latest Columns",   df2.shape[1])

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
st.markdown(f"### Step 3 — QC Results  •  {REPORTS[rtype]['icon']} {rtype}")

# ══════════════════════════════════════════════════════════════════════════════
# RUN CHECKS
# ══════════════════════════════════════════════════════════════════════════════

checks = []  # each: {name, ok, msg, df}

with st.spinner("Running checks…"):

    # ── OAM ──────────────────────────────────────────────────────────────────
    if rtype == "OAM":

        # Check 1 – Empty columns
        for label, df in [("Previous", df1), ("Latest", df2)]:
            empty = [c for c in df.columns if blank(df[c]).all()]
            ok = len(empty) == 0
            r = {"name": f"1. Empty Columns ({label})", "ok": ok,
                 "msg": f"{len(empty)} completely empty column(s) in {label}",
                 "df": pd.DataFrame({"Empty Column": empty}) if not ok else None}
            checks.append(r)
            show_check("1", f"Empty Columns ({label})", r["ok"], r["msg"], r["df"])

        # Check 2 – PMC without DDD
        for label, df in [("Previous", df1), ("Latest", df2)]:
            if "PMC_ACCT_ID" in df.columns and "OUTLET_DDD" in df.columns:
                sub = df[~blank(df["PMC_ACCT_ID"])]
                bad = sub[blank(sub["OUTLET_DDD"])][["PMC_ACCT_ID"]].drop_duplicates()
                ok = len(bad) == 0
                r = {"name": f"2. PMC without DDD ({label})", "ok": ok,
                     "msg": f"{len(bad)} PMC_ACCT_ID(s) missing DDD in {label}",
                     "df": bad if not ok else None}
                checks.append(r)
                show_check("2", f"PMC_ACCT_ID Without DDD ({label})", r["ok"], r["msg"], r["df"])

        # Check 3 – Blank TYP
        for label, df in [("Previous", df1), ("Latest", df2)]:
            if "TYP" in df.columns:
                bad = df[blank(df["TYP"])].reset_index().rename(columns={"index": "row"})
                ok = len(bad) == 0
                r = {"name": f"3. Blank TYP ({label})", "ok": ok,
                     "msg": f"{len(bad)} blank TYP row(s) in {label}",
                     "df": bad[["row", "TYP"]].head(200) if not ok else None}
                checks.append(r)
                show_check("3", f"Blank TYP ({label})", r["ok"], r["msg"], r["df"])

        # Check 4 – PMC mapped to multiple DDDs
        for label, df in [("Previous", df1), ("Latest", df2)]:
            if "PMC_ACCT_ID" in df.columns and "OUTLET_DDD" in df.columns:
                sub = df[~blank(df["PMC_ACCT_ID"]) & ~blank(df["OUTLET_DDD"])]
                multi = (sub.groupby("PMC_ACCT_ID")["OUTLET_DDD"]
                           .agg(lambda x: " | ".join(sorted(x.unique())))
                           .reset_index(name="DDDs"))
                multi = multi[multi["DDDs"].str.contains(r"\|", regex=True)]
                ok = len(multi) == 0
                r = {"name": f"4. PMC Multi-DDD ({label})", "ok": ok,
                     "msg": f"{len(multi)} PMC_ACCT_ID(s) mapped to multiple DDDs in {label}",
                     "df": multi if not ok else None}
                checks.append(r)
                show_check("4", f"PMC → Multiple DDDs ({label})", r["ok"], r["msg"], r["df"])

        # Check 5 – Parent/Child change per DDD
        ok5, msg5, df5 = soam_check5_child_parent(df1, df2)
        checks.append({"name": "5. Parent/Child Change per DDD", "ok": ok5, "msg": msg5, "df": df5})
        show_check("5", "Parent / Child ID & Name Changes per DDD", ok5, msg5, df5)

        # Check 6 – DDD counts
        c1_ddd = df1["OUTLET_DDD"].value_counts() if "OUTLET_DDD" in df1.columns else pd.Series(dtype=int)
        c2_ddd = df2["OUTLET_DDD"].value_counts() if "OUTLET_DDD" in df2.columns else pd.Series(dtype=int)
        cnt_rows = []
        for d in c1_ddd.index.union(c2_ddd.index):
            v1, v2 = c1_ddd.get(d, 0), c2_ddd.get(d, 0)
            if v1 != v2:
                cnt_rows.append({"OUTLET_DDD": d, "Count Previous": int(v1),
                                 "Count Latest": int(v2), "Delta": int(v2 - v1)})
        summary6_rows = []
        for label, df in [("Previous", df1), ("Latest", df2)]:
            if "OUTLET_DDD" in df.columns:
                summary6_rows.append({"Version": label,
                    "Total rows with DDD": int(df["OUTLET_DDD"].notna().sum()),
                    "Distinct DDDs": int(df["OUTLET_DDD"].nunique())})
        cnt_df = pd.DataFrame(cnt_rows)
        ok6 = len(cnt_df) == 0
        r6 = {"name": "6. DDD Counts & Distinct", "ok": ok6,
              "msg": f"{len(cnt_df)} DDD(s) with row-count change",
              "df": cnt_df if not ok6 else None}
        checks.append(r6)
        with st.expander(f"**6. DDD Counts & Distinct**  {rb(ok6)}", expanded=not ok6):
            st.dataframe(pd.DataFrame(summary6_rows), use_container_width=True, hide_index=True)
            if not ok6:
                st.markdown("**DDDs with count changes:**")
                st.dataframe(cnt_df, use_container_width=True, hide_index=True)

        # Check 7 – NPI → one DDD
        for label, df in [("Previous", df1), ("Latest", df2)]:
            if "NPI" in df.columns and "OUTLET_DDD" in df.columns:
                sub = df[~blank(df["NPI"]) & ~blank(df["OUTLET_DDD"])]
                multi = (sub.groupby("NPI")["OUTLET_DDD"]
                           .agg(lambda x: " | ".join(sorted(x.unique())))
                           .reset_index(name="DDDs"))
                multi = multi[multi["DDDs"].str.contains(r"\|", regex=True)]
                ok = len(multi) == 0
                r = {"name": f"7. NPI Multi-DDD ({label})", "ok": ok,
                     "msg": f"{len(multi)} NPI(s) mapped to multiple DDDs in {label}",
                     "df": multi if not ok else None}
                checks.append(r)
                show_check("7", f"NPI → Multiple DDDs ({label})", r["ok"], r["msg"], r["df"])

        # Check 8 – ID_867 → one DDD
        for label, df in [("Previous", df1), ("Latest", df2)]:
            if "ID_867" in df.columns and "OUTLET_DDD" in df.columns:
                sub = df[~blank(df["ID_867"]) & ~blank(df["OUTLET_DDD"])]
                multi = (sub.groupby("ID_867")["OUTLET_DDD"]
                           .agg(lambda x: " | ".join(sorted(x.unique())))
                           .reset_index(name="DDDs"))
                multi = multi[multi["DDDs"].str.contains(r"\|", regex=True)]
                ok = len(multi) == 0
                r = {"name": f"8. ID_867 Multi-DDD ({label})", "ok": ok,
                     "msg": f"{len(multi)} ID_867(s) mapped to multiple DDDs in {label}",
                     "df": multi if not ok else None}
                checks.append(r)
                show_check("8", f"ID_867 → Multiple DDDs ({label})", r["ok"], r["msg"], r["df"])

    # ── SUPER OAM ─────────────────────────────────────────────────────────────
    elif rtype == "Super OAM":
        r = soam_check1_ddd_id867(df1, df2)
        checks.append({"name": "1. DDD change per ID_867", "ok": r[0], "msg": r[1], "df": r[2]})
        show_check("1", "DDD Change per ID_867", r[0], r[1], r[2])

        r = soam_check2_exclusion_flags(df1, df2)
        checks.append({"name": "2. Exclusion Flag Changes per DDD", "ok": r[0], "msg": r[1], "df": r[2]})
        show_check("2", "Exclusion Flags (Subcat / PED / Override) per DDD", r[0], r[1], r[2])

        r = soam_check3_territory_unaligned(df1, df2)
        checks.append({"name": "3. Territory Column — Unaligned", "ok": r[0], "msg": r[1], "df": r[2]})
        show_check("3", "Territory Columns — Unaligned check", r[0], r[1], r[2])

        r = soam_check4_smart_flags(df1, df2)
        checks.append({"name": "4. SMART Flag Distribution", "ok": r[0], "msg": r[1], "df": r[2]})
        show_check("4", "SMART Flag Comparison (All Brands)", r[0], r[1], r[2])

        r = soam_check5_child_parent(df1, df2)
        checks.append({"name": "5. Child / Parent Hierarchy per DDD", "ok": r[0], "msg": r[1], "df": r[2]})
        show_check("5", "Child / Parent ID & Name per DDD", r[0], r[1], r[2])

        r = soam_check6_ic_channels(df1, df2)
        checks.append({"name": "6. IC Channel Flag Distribution", "ok": r[0], "msg": r[1], "df": r[2]})
        show_check("6", "IC Channel Flags Comparison (All Brands)", r[0], r[1], r[2])

        r = soam_check7_monthly_cols(df1, df2)
        checks.append({"name": "7. Monthly (yyyymm) Column Comparison", "ok": r[0], "msg": r[1], "df": r[2]})
        show_check("7", "Monthly (yyyymm) Column Presence & Sum Diff", r[0], r[1], r[2])

    # ── NPS FEEDS / IC FEEDS ──────────────────────────────────────────────────
    elif rtype in ("NPS Feeds", "IC Feeds"):
        r = feed_check1_schema(df1, df2)
        checks.append({"name": "1. Schema (Column Count / Length)", "ok": r[0], "msg": r[1], "df": r[2]})
        show_check("1", "Schema — Column Count & Max Length", r[0], r[1], r[2])

        # Check 2 — Column Type, Length & Blank Cross-check
        r5 = feed_check5_col_type_length(df1, df2)
        checks.append({"name": "2. Column Type, Length & Blank Check", "ok": r5[0], "msg": r5[1], "df": r5[2]})
        with st.expander(f"**2. Column Type, Length & Blank Check**  {rb(r5[0])}", expanded=(r5[0] is False)):
            st.markdown(f"{'✅' if r5[0] else '❌'} {r5[1]}")
            if r5[2] is not None and len(r5[2]):
                view = st.radio("Show columns:", ["Issues only", "All columns"],
                                horizontal=True, key="chk2_view")
                disp = r5[2][r5[2]["Issue"] != "OK"] if view == "Issues only" else r5[2]
                st.dataframe(disp, use_container_width=True, hide_index=True)

        # Check 3 — Row Counts
        r6 = feed_check6_row_counts(df1, df2)
        checks.append({"name": "3. Row Count Comparison", "ok": r6[0], "msg": r6[1], "df": r6[3]})
        with st.expander(f"**3. Row Count Comparison**  {rb(r6[0])}", expanded=(r6[0] is False)):
            st.markdown(f"{'✅' if r6[0] else '❌'} {r6[1]}")
            st.dataframe(r6[2], use_container_width=True, hide_index=True)
            if r6[3] is not None and len(r6[3]):
                st.markdown("**Columns with non-blank row count change:**")
                st.dataframe(r6[3], use_container_width=True, hide_index=True)

        # Check 4 — IC Feeds only: TotalRx vs TotalQty
        if rtype == "IC Feeds":
            r7 = feed_check7_ic_totalrx_qty(df1, df2)
            checks.append({"name": "4. TotalRx vs TotalQty Row Counts", "ok": r7[0], "msg": r7[1], "df": r7[2]})
            show_check("4", "TotalRx & TotalQty — Row Count Comparison (IC Feeds)", r7[0], r7[1], r7[2])
            chk_offset = 5
        else:
            chk_offset = 4

        r = feed_check3_monthly(df1, df2)
        checks.append({"name": f"{chk_offset}. Monthly HCP / Account Comparison", "ok": r[0], "msg": r[1], "df": r[2]})
        show_check(str(chk_offset), "Monthly Comparison — HCP / Account Additions & Removals", r[0], r[1], r[2])

        # ── Check 4 — Alignment ID Monthly Comparison ────────────────────────
        st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
        _align_chk_num = chk_offset + 1
        st.markdown(f"#### ⚙️ Check {_align_chk_num} — Alignment ID Monthly Comparison (column mapping)")
        st.caption("Auto-detection is shown below. Override any column if needed, then click **Run Check 4**.")

        all_cols_feed = list(df1.columns)

        # Auto-detect date column
        def _guess_date(df):
            for c in df.columns:
                s = df[c].dropna().astype(str).str.strip().head(100)
                if s.str.fullmatch(r"\d{6}").mean() > 0.5:
                    return c
                try:
                    pd.to_datetime(s, errors="raise")
                    return c
                except Exception:
                    pass
            return all_cols_feed[0]

        # Auto-detect alignment ID column
        def _guess_align(df):
            for c in df.columns:
                cu = c.upper().replace(" ", "_")
                if "ALIGN" in cu and ("ID" in cu or "KEY" in cu):
                    return c
            for c in df.columns:
                if "ALIGN" in c.upper():
                    return c
            return all_cols_feed[0]

        # Auto-detect value column
        def _guess_value(df, rtype):
            kw = "TOTALRX" if rtype == "IC Feeds" else "NPS"
            for c in df.columns:
                if kw in c.upper().replace(" ", "").replace("_", ""):
                    return c
            # fallback: first numeric-looking column
            for c in df.columns:
                if pd.to_numeric(df[c], errors="coerce").notna().sum() > len(df) * 0.3:
                    return c
            return all_cols_feed[0]

        default_date  = _guess_date(df1)
        default_align = _guess_align(df1)
        default_value = _guess_value(df1, rtype)
        value_label   = "TotalRX column" if rtype == "IC Feeds" else "NPS column"

        col_d, col_a, col_v = st.columns(3)
        with col_d:
            sel_date = st.selectbox("📅 Date column (yyyymm extracted from this)",
                                    options=all_cols_feed,
                                    index=all_cols_feed.index(default_date) if default_date in all_cols_feed else 0,
                                    key="sel_date")
        with col_a:
            sel_align = st.selectbox("🔑 Alignment ID column",
                                     options=all_cols_feed,
                                     index=all_cols_feed.index(default_align) if default_align in all_cols_feed else 0,
                                     key="sel_align")
        with col_v:
            sel_value = st.selectbox(f"📊 {value_label}",
                                     options=all_cols_feed,
                                     index=all_cols_feed.index(default_value) if default_value in all_cols_feed else 0,
                                     key="sel_value")

        if st.button(f"▶️  Run Check {_align_chk_num} — Alignment Monthly Comparison",
                     use_container_width=True, type="primary", key="run_chk4"):
            st.session_state["chk4_run"]   = True
            st.session_state["chk4_date"]  = sel_date
            st.session_state["chk4_align"] = sel_align
            st.session_state["chk4_value"] = sel_value

        if st.session_state.get("chk4_run"):
            _d = st.session_state["chk4_date"]
            _a = st.session_state["chk4_align"]
            _v = st.session_state["chk4_value"]
            with st.spinner("Running alignment monthly comparison…"):
                try:
                    ok4, msg4, period_df, detail_df = feed_check4_alignment_monthly(
                        df1, df2, _d, _a, _v)
                    checks.append({"name": f"{_align_chk_num}. Alignment ID Monthly Comparison",
                                   "ok": ok4, "msg": msg4, "df": detail_df})
                    checks.append({"name": f"{_align_chk_num}a. Monthly Period Summary",
                                   "ok": ok4, "msg": msg4, "df": period_df})
                    with st.expander(f"**{_align_chk_num}. Alignment ID Monthly Comparison**  {rb(ok4)}",
                                     expanded=True):
                        st.markdown(f"{'✅' if ok4 else '❌'} {msg4}")

                        st.markdown("##### 📅 Period-level Summary")
                        st.dataframe(period_df, use_container_width=True, hide_index=True)

                        st.markdown("##### 🔍 Full Alignment × Period Detail")
                        status_filter = st.multiselect(
                            "Filter by status",
                            options=["Matched", "Changed", "New in Latest", "Removed in Latest"],
                            default=["Changed", "New in Latest", "Removed in Latest"],
                            key="chk4_filter")
                        filtered = detail_df[detail_df["Status"].isin(status_filter)] if status_filter else detail_df
                        st.dataframe(filtered, use_container_width=True, hide_index=True)
                        st.caption(f"Showing {len(filtered):,} of {len(detail_df):,} rows")
                except Exception as e:
                    st.error(f"Check 4 failed: {e}")

# ══════════════════════════════════════════════════════════════════════════════
# EXPORT
# ══════════════════════════════════════════════════════════════════════════════

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
st.markdown("### 📥 Save Report")
st.caption("The Excel report contains: ✅ QC Checklist (manual sign-off), 📋 Summary, and one detail sheet per finding.")

with st.spinner("Building Excel…"):
    try:
        xlsx = build_excel(rtype, checks)
        build_ok = True
        build_err = ""
    except Exception as _e:
        build_ok = False
        build_err = str(_e)
        xlsx = b""

if not build_ok:
    st.error(f"Excel build failed: {build_err}")
else:
    ts_str   = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname    = f"{rtype.replace(' ', '_')}_QC_{ts_str}.xlsx"

    st.download_button(
        label=f"⬇️  Download {rtype} QC Report (.xlsx)",
        data=xlsx,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
